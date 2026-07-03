"""
Export Layer — Excel (.xlsx)
------------------------------
Generates a formatted Excel workbook with one row per invoice, plus a second
sheet with normalized line items. Useful for accounting hand-off or as a
downloadable snapshot from the review UI.
"""
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config import settings
from app.database import Invoice


def _build_invoice_summary(inv: Invoice) -> str:
    """
    Builds a concise one-line description of what an invoice is about,
    derived entirely from already-extracted fields — no extra LLM call.
    """
    parts = []

    # Vendor → Customer direction
    if inv.vendor_name and inv.customer_name:
        parts.append(f"{inv.vendor_name} → {inv.customer_name}")
    elif inv.vendor_name:
        parts.append(f"From {inv.vendor_name}")
    elif inv.customer_name:
        parts.append(f"To {inv.customer_name}")

    # What was purchased — use up to 2 line item descriptions
    if inv.line_items_json:
        try:
            items = json.loads(inv.line_items_json)
            descriptions = [
                str(i.get("description", "")).strip()
                for i in items
                if i.get("description")
            ]
            if descriptions:
                shown = descriptions[:2]
                suffix = f" +{len(descriptions) - 2} more" if len(descriptions) > 2 else ""
                parts.append(f"for {', '.join(shown)}{suffix}")
        except (json.JSONDecodeError, TypeError):
            pass

    # Total amount
    if inv.total_amount is not None:
        currency = inv.currency or ""
        parts.append(f"totalling {currency} {inv.total_amount:,.2f}".strip())

    # Payment terms
    if inv.payment_terms:
        parts.append(f"({inv.payment_terms})")

    return "; ".join(parts) if parts else "—"


def export_invoices_to_excel(db: Session, invoice_ids: list[str] | None = None) -> str:
    query = db.query(Invoice)
    if invoice_ids:
        query = query.filter(Invoice.id.in_(invoice_ids))
    invoices = query.all()

    wb = Workbook()
    ws_inv = wb.active
    ws_inv.title = "Invoices"
    ws_items = wb.create_sheet("Line Items")

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    currency_style = NamedStyle(name="currency")
    currency_style.number_format = '#,##0.00'
    
    # Define headers
    INV_HEADERS = [
        "Invoice Number", "Invoice Date", "Due Date",
        "Vendor Name", "Customer Name", "Currency", "Subtotal",
        "Tax Amount", "Total Amount", "Status", "Details"
    ]
    
    ITEM_HEADERS = [
        "Invoice Number", "Description", "Quantity", 
        "Unit Price", "Line Total"
    ]

    ws_inv.append(INV_HEADERS)
    ws_items.append(ITEM_HEADERS)
    
    # Format Headers
    for ws in (ws_inv, ws_items):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # Data Rows
    for inv in invoices:
        ws_inv.append([
            inv.invoice_number,
            inv.invoice_date,
            inv.due_date,
            inv.vendor_name,
            inv.customer_name,
            inv.currency,
            inv.subtotal,
            inv.tax_amount,
            inv.total_amount,
            inv.status.replace("_", " ").title() if inv.status else "",
            inv.invoice_summary or _build_invoice_summary(inv),
        ])
        
        # Apply currency format to subtotal, tax, total
        row_idx = ws_inv.max_row
        for col_idx in [7, 8, 9]:  # Subtotal, Tax Amount, Total Amount
            ws_inv.cell(row=row_idx, column=col_idx).style = currency_style
        
        if inv.line_items_json:
            try:
                items = json.loads(inv.line_items_json)
            except json.JSONDecodeError:
                items = []
            for item in items:
                ws_items.append([
                    inv.invoice_number,
                    item.get("description"), 
                    item.get("quantity"),
                    item.get("unit_price"), 
                    item.get("line_total")
                ])
                # Apply currency format to unit price and line total
                item_row_idx = ws_items.max_row
                for col_idx in [4, 5]:
                    ws_items.cell(row=item_row_idx, column=col_idx).style = currency_style

    # Auto-adjust column widths
    for ws in (ws_inv, ws_items):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 4, 40) # Max width of 40
            ws.column_dimensions[col_letter].width = adjusted_width

    # Format as Excel Tables
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    if len(invoices) > 0:
        tab_inv = Table(displayName="InvoicesTable", ref=f"A1:{get_column_letter(len(INV_HEADERS))}{len(invoices)+1}")
        style_inv = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab_inv.tableStyleInfo = style_inv
        ws_inv.add_table(tab_inv)
        
        if ws_items.max_row > 1:
            tab_items = Table(displayName="LineItemsTable", ref=f"A1:{get_column_letter(len(ITEM_HEADERS))}{ws_items.max_row}")
            tab_items.tableStyleInfo = style_inv
            ws_items.add_table(tab_items)

    os.makedirs(settings.excel_export_dir, exist_ok=True)
    filename = f"invoices_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(settings.excel_export_dir, filename)
    wb.save(out_path)
    return out_path
